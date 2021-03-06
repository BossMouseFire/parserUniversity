import json
import openpyxl


FILEname = "/Users/danil/PycharmProjects/skills_university/Statistic information Test.xlsx"

try:
    wb = openpyxl.load_workbook(FILEname)
except Exception as error:
    wb = openpyxl.Workbook()

    # Удаление листа, создаваемого по умолчанию, при создании документа
    for sheet_name in wb.sheetnames:
        sheet = wb.get_sheet_by_name(sheet_name)
        wb.remove_sheet(sheet)
    firstlist = wb.create_sheet('Statistics')
    secondlist = wb.create_sheet('Full Statistics')
firstlist = wb.get_sheet_by_name('Statistics')
secondlist = wb.get_sheet_by_name('Full Statistics')


def sheetdecoration():
    firstlist['A1'] = "Фамилия Имя ученого"
    firstlist['B1'] = "Количество ключевых слов на reseachgate"
    firstlist['C1'] = "Количество ключевых слов на elibrary"
    firstlist['D1'] = "Количество полных совпадений"
    firstlist['E1'] = "Количество хороших совпадений"
    firstlist['F1'] = "Количество возможных совпадений"
    firstlist['G1'] = "Вывод"

    secondlist['A1'] = "Фамилия Имя ученого"
    secondlist['B1'] = "Список ключевых слов"
    secondlist['D1'] = "Количество совпадений"
    secondlist['E1'] = "Список совпавших слов"
    secondlist['F1'] = "Список возможных совпадений"
    secondlist['H1'] = "Количество возможных совпадений"

    secondlist['B2'] = "ResearchGate"
    secondlist['C2'] = "Elibrary"
    secondlist['F2'] = "ResearchGate"
    secondlist['G2'] = "Elibrary"

    secondlist.merge_cells('B1:C1')
    secondlist.merge_cells('F1:G1')
    secondlist.merge_cells('A1:A2')
    secondlist.merge_cells('D1:D2')
    secondlist.merge_cells('E1:E2')
    secondlist.merge_cells('H1:H2')
    row_cells = 1
    while row_cells < 169:
        firstlist.row_dimensions[row_cells].height = 15
        row_cells += 1
    column_cells = 65
    while column_cells < 75:
        firstlist.column_dimensions[chr(column_cells)].width = 40
        secondlist.column_dimensions[chr(column_cells)].width = 30
        column_cells += 1


sheetdecoration()

with open("profiles_researchgate.json", "r", encoding="utf8") as file:
    dataElibrary = json.load(file)
with open("profiles_elibrary.json", "r", encoding="utf8") as file:
    dataResearch = json.load(file)

Sovpadenia = 0
PossibleSovpadenia = 0
PossibleSovpadeniaHuman = 0
amountoffullsovpad = {}  # количество полных совпадений
goodsovpad = {}  # количество хороших совпадений (2 корня)
anysovpad = {}  # количество возможных совпадений
columsADEH = 3
# "Умное" сравнивание
NotFindNamesSum = 0
NotFindNamesSum2 = 0
row_number_name = 2

# группы
average = 0
one_group = 0
two_group = 0
three_group = 0
four_group = 0

for NameElibrary in dataElibrary:  # начинаем поиск с elibrary
    NotFindName = True
    firstlist.cell(row=row_number_name, column=1).value = NameElibrary
    firstlist.cell(row=row_number_name, column=2).value = len(dataElibrary[NameElibrary])
    try:
        firstlist.cell(row=row_number_name, column=3).value = len(dataResearch[NameElibrary])
    except KeyError:
        firstlist.cell(row=row_number_name, column=3).value = "Имя не найдено"
    for NameResearch in dataResearch:  # ищем совпадающее имя на research
        if NameElibrary == NameResearch:  # имя нашлось
            NotFindName = False
            for keyWordElibrary in dataElibrary[NameElibrary]:  # выбираем ключевое слово из имени Elibrary
                INOTFIND = True
                for keyWordResearch in dataResearch[NameResearch]: # пытаемся найти полное совпадение ключего слова на research
                    if keyWordElibrary == keyWordResearch:
                        # print(key2)
                        Sovpadenia += 1
                        for namei in amountoffullsovpad.keys():
                            if namei == NameElibrary:
                                amountoffullsovpad[namei] += 1
                                break
                        else:
                            amountoffullsovpad.update({NameElibrary: 1})
                        Del = dataResearch[NameResearch].index(keyWordResearch)
                        dataResearch[NameResearch].pop(Del)

                        Del = dataElibrary[NameElibrary].index(keyWordElibrary)
                        dataElibrary[NameElibrary].pop(Del)
                        INOTFIND = False
                        break
                if INOTFIND:  # если не было найдено ключевое слово, начинаем поиск совпадений
                    FINDPOSSIBLE = False

                    for keyWordResearch in dataResearch[NameResearch]:
                        if len(keyWordResearch) > 5:
                            tempstr1 = str(keyWordElibrary).split(' ')  # список из слов ключевого набора Research
                            tempstr2 = str(keyWordResearch).split(' ')  # 2 список из слов ключевого набора 2
                            FINDwithROOT = 0
                            ListofROOTS = {}

                            for keySplitResearch in tempstr1:  # слова из списка ключегого набора
                                TempDelletter = 0
                                TempLen = len(keySplitResearch)  # длина слова
                                ROOTUSE = False
                                while TempLen + TempDelletter > 5:
                                    if ROOTUSE:
                                        break
                                    if TempDelletter == 0:
                                        RootStr = keySplitResearch
                                    else:
                                        RootStr = str(keySplitResearch)[0:TempDelletter]  # "КОРЕНЬ" слова
                                    DifferencefromROOT = 0

                                    # "Умное" сравнивание
                                    for keySplitElibrary in tempstr2:
                                        if ROOTUSE:
                                            break
                                        letter = 0
                                        lengthWord = len(keySplitElibrary)
                                        while len(keySplitElibrary) + letter >= len(RootStr):  # Проверяем на равенство "корня" и слова из 2 списка
                                            if letter == 0:
                                                if keySplitElibrary == RootStr:
                                                    FINDPOSSIBLE = True
                                                    ROOTUSE = True
                                                    FINDwithROOT += 1
                                                    DifferencefromROOT = abs(letter)
                                                    ListofROOTS.update({RootStr: DifferencefromROOT})
                                                    break
                                            elif str(keySplitElibrary)[0:letter] == RootStr:
                                                FINDPOSSIBLE = True
                                                ROOTUSE = True
                                                FINDwithROOT += 1
                                                DifferencefromROOT = abs(letter)
                                                ListofROOTS.update({RootStr: DifferencefromROOT})
                                                break
                                            letter -= 1
                                    TempDelletter -= 1
                            if FINDwithROOT > 0:
                                if FINDwithROOT > 1:
                                    for namei in goodsovpad.keys():
                                        if namei == NameElibrary:
                                            goodsovpad[namei] += 1
                                            break
                                    else:
                                        goodsovpad.update({NameElibrary: 1})
                                print(NameElibrary)  # имя человека
                                print(tempstr1)  # строка research
                                print(tempstr2)  # строка elibrary
                                print(ListofROOTS)
                                print(FINDwithROOT)
                                print('\n')
                    if FINDPOSSIBLE:
                        PossibleSovpadenia += 1
    try:
        firstlist.cell(row=row_number_name, column=4).value = amountoffullsovpad[NameElibrary]
    except KeyError:
        firstlist.cell(row=row_number_name, column=4).value = 0
        amountoffullsovpad.update({
            NameElibrary: 0
        })
    try:
        firstlist.cell(row=row_number_name, column=5).value = goodsovpad[NameElibrary]
    except KeyError:
        firstlist.cell(row=row_number_name, column=5).value = 0
        goodsovpad.update({
            NameElibrary: 0
        })
    numAnySov = PossibleSovpadenia - PossibleSovpadeniaHuman
    anysovpad.update({
        NameElibrary: numAnySov
    })
    PossibleSovpadeniaHuman = PossibleSovpadenia
    firstlist.cell(row=row_number_name, column=6).value = anysovpad[NameElibrary]
    try:
        quality = (amountoffullsovpad[NameElibrary] + goodsovpad[NameElibrary] * 0.25 +
                   anysovpad[NameElibrary] * 0.1) * 100 / \
                  (len(dataElibrary[NameElibrary]) + amountoffullsovpad[NameElibrary])
        average += quality
        if quality > 50:
            one_group += 1
        elif quality > 40:
            two_group += 1
        elif quality > 25:
            three_group += 1
        else:
            four_group += 1
        firstlist.cell(row=row_number_name, column=7).value = f'Процент совпадения elibrary c ' \
            f'researchgate равен {round(quality, 2)}%.'
    except ZeroDivisionError:
        firstlist.cell(row=row_number_name, column=7).value = "Машинное обучение не определило ключевые слова"
    if NotFindName:
        NotFindNamesSum += 1
        print(NameElibrary + "   NOT FIND NAME")
    else:
        NotFindNamesSum2 += 1
    row_number_name += 1


wb.save(FILEname)


print('Количество ненайденных имен: ' + str(NotFindNamesSum))
print('Количество найденных имен: ' + str(NotFindNamesSum2))
print('Полные совпадения слов: ' + str(Sovpadenia))
print('Возможные совпадения слов: ' + str(PossibleSovpadenia))
print('\n')
print("Количество полных совпадений у ученых: " + str(amountoffullsovpad))
print("Количество хороших совпадений у ученых (у ключевых слов обнаружено совпадение двух и более подслов): " + str(goodsovpad))
print('Среднее', average / 110)
print(one_group, two_group, three_group, four_group)
